import openpyxl
import os
os.chdir(os.path.dirname(__file__))


# Reads Data from Excel
def readData(sheetName, rowNum, columnNum, file=os.getcwd()+"/data.xlsx"):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=columnNum).value


# Writes Data to the Excel
def writeData(sheetName, rowNum, columnNum, data, file=os.getcwd()+"/data.xlsx"):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=columnNum).value = data
    workbook.save(file)
